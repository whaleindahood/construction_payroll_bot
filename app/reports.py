from __future__ import annotations

import csv
import re
from collections.abc import Sequence
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def spreadsheet_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def table_csv(headers: Sequence[str], rows: Sequence[Sequence[object]]) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(headers)
    writer.writerows([[spreadsheet_safe(value) for value in row] for row in rows])
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def table_xlsx(title: str, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> bytes:
    book = Workbook()
    sheet = book.active
    if sheet is None:
        raise RuntimeError("Не удалось создать лист Excel.")
    sheet.title = re.sub(r"[\\/*?:\[\]]", "_", title)[:31] or "Report"
    sheet.append(list(headers))
    for row in rows:
        sheet.append([spreadsheet_safe(value) for value in row])
    sheet.freeze_panes = "A2"
    for index, column in enumerate(sheet.columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            max(len(str(cell.value or "")) for cell in column) + 2
        )
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()
