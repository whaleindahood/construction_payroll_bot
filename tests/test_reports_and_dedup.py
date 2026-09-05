from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from app.bot.handlers import totals_by_currency
from app.reports import summaries_csv, summaries_xlsx, table_csv, table_xlsx
from app.services import UpdateDedupService
from tests.test_financial_invariants import ACTOR, setup_employee_object


def test_object_report_and_exports(services):
    employee, obj = setup_employee_object(services)
    services["attendance"].create_bulk(
        employee_ids=[employee.id],
        object_id=obj.id,
        work_date=date(2026, 9, 4),
        coefficient="0.5",
        actor=ACTOR,
        operation_key="report-day",
    )
    report_object, rows = services["reports"].object_report(
        obj.id, date(2026, 9, 1), date(2026, 9, 30)
    )
    summaries = services["payroll"].all_summaries(
        date_from=date(2026, 9, 1), date_to=date(2026, 9, 30)
    )

    assert report_object.id == obj.id
    assert rows[0][0] == employee.name
    assert str(rows[0][2]) == "0.50"
    csv_data = summaries_csv(summaries)
    assert employee.name.encode("utf-8") in csv_data
    workbook = load_workbook(filename=__import__("io").BytesIO(summaries_xlsx(summaries)))
    assert workbook.active["A2"].value == employee.name


def test_telegram_update_claim_is_idempotent(services):
    dedup = UpdateDedupService(services["sessions"])

    assert dedup.claim(77) is True
    assert dedup.claim(77) is False
    dedup.done(77)
    assert dedup.claim(77) is False


def test_exports_neutralize_spreadsheet_formulas():
    csv_data = table_csv(["Name"], [['=HYPERLINK("bad")']]).decode("utf-8")
    workbook = load_workbook(
        filename=__import__("io").BytesIO(table_xlsx("Bad/Title", ["Name"], [["+cmd"]]))
    )

    assert "'=HYPERLINK" in csv_data
    assert workbook.active.title == "Bad_Title"
    assert workbook.active["A2"].value == "'+cmd"


def test_mixed_currency_totals_stay_separate(services):
    first, _ = setup_employee_object(services)
    second = services["employees"].create(
        name="Рублёвый сотрудник",
        rate="1000",
        currency="RUB",
        start_date=date(2026, 9, 1),
        actor=ACTOR,
    )

    totals = totals_by_currency(
        [services["payroll"].summary(first.id), services["payroll"].summary(second.id)]
    )

    assert set(totals) == {"EUR", "RUB"}
